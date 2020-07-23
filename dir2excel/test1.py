#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2020/7/23 4:17 下午
# @Author  : silianpan
# @Site    : 
# @File    : test1.py
# @Software: PyCharm
from openpyxl import load_workbook
from dir2excel import get_file_list


def start():
    # 写数据
    newfile = './ret.xlsx'
    wb = load_workbook(newfile)
    ws = wb['Sheet1']
    # 直接根据位置进行赋值
    ws['A1'] = '大类'
    ws['B1'] = '类名'
    ws['C1'] = '说明'

    source_dir = '/Users/liupan/work/code/power-plus/code/backend/power-plus-server/src/main/java'
    ret = []
    get_file_list(source_dir, ret, source_dir)

    i = 2
    for item in ret:
        ws.cell(row=i, column=2, value=str(item))
        print(item)
        i = i + 1

    wb.save(newfile)


if __name__ == '__main__':
    start()
