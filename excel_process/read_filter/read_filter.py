#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2021/3/3 下午4:22
# @Author  : silianpan
# @Site    : 
# @File    : read_filter.py
# @Software: PyCharm
# @description 本工具主要采用openpyxl插件，openpyxl插件主要用于读写Excel

import openpyxl


# 读取关键词文本key.txt
# 采用open方法，循环读取每一行，并去掉换行符\n
def read_keys():
    key_list = []
    with open('./key.txt', 'r') as f:
        for line in f.readlines():
            key_list.append(line.strip('\n'))
    return key_list


# 读取excel并过滤
def read_excel():
    keys = read_keys();
    excel_file = r'./test.xlsx'
    inwb = openpyxl.load_workbook(excel_file)
    ws = inwb['Sheet1']

    # 获取sheet的最大行数和列数
    rows = ws.max_row
    cols = ws.max_column
    all_item = []
    for r in range(2, rows + 1):
        name = ws.cell(r, 1).value
        cust_id_type = ws.cell(r, 2).value
        cust_id_no = ws.cell(r, 3).value
        remark = ws.cell(r, 4).value
        # 过滤remark
        # 循环关键词
        for key in keys:
            # 如果remark包含关键词，找出这条记录
            if key in remark:
                item = {
                    'name': name,
                    'cust_id_type': cust_id_type,
                    'cust_id_no': cust_id_no,
                    'remark': remark
                }
                all_item.append(item)
                # 跳出循环，不在遍历其他关键词，继续进行下一行的检测
                break
    # 返回过滤后的结果
    return all_item

# 将数据写入到新的excel表中
def write_excel():
    # 注意：这里要先新建空ret.xlsx文件
    newfile = './ret.xlsx'
    wb = openpyxl.load_workbook(newfile)
    ws = wb['Sheet1']
    # 直接根据位置进行赋值
    ws['A1'] = 'NAME'
    ws['B1'] = 'CUST_ID_TYPE'
    ws['C1'] = 'CUST_ID_NO'
    ws['D1'] = 'REMARK'

    # 读excel获取结果
    all_item = read_excel()
    # 从第二行第一列开始写数据，第一行已经作为标题
    i = 2
    for item in all_item:
        ws.cell(row=i, column=1, value=str(item))
        i = i + 1
    wb.save(newfile)


# 程序入口
if __name__ == '__main__':
    write_excel()
