#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/1/5 下午4:25
# @Author  : silianpan
# @Site    :
# @File    : test.py
# @Software: PyCharm
# @description 排列组合求均值，本工具主要采用openpyxl插件，openpyxl插件主要用于读写Excel

import openpyxl
import itertools


# 读取excel并过滤
def read_excel():
    excel_file = r"./data.xlsx"
    inwb = openpyxl.load_workbook(excel_file)
    ws = inwb["3混1"]

    # 获取sheet的最大行数和列数
    # rows = ws.max_row
    cols = ws.max_column
    all_item = []
    for r in range(3, cols + 1):
        # name = ws.cell(1, r).value
        d1 = ws.cell(2, r).value
        d2 = ws.cell(3, r).value
        d3 = ws.cell(4, r).value
        d4 = ws.cell(5, r).value
        d5 = ws.cell(6, r).value
        d9 = ws.cell(7, r).value
        d10 = ws.cell(8, r).value
        d11 = ws.cell(9, r).value
        d12 = ws.cell(10, r).value
        d13 = ws.cell(11, r).value
        all_item.append((d1, d2, d3, d4, d5, d9, d10, d11, d12, d13))
    print(all_item)
    return all_item


# 将数据写入到新的excel表中
def write_excel(no):
    # 注意：这里要先新建空ret.xlsx文件
    newfile = "./data.xlsx"
    wb = openpyxl.load_workbook(newfile)
    ws = wb[str(no) + "混1"]

    # 读excel获取结果
    all_item = read_excel()

    numbers = {"1", "2", "3", "4", "5", "9", "10", "11", "12", "13"}
    tmp_i = 1
    for i in itertools.combinations(numbers, no):
        ws.cell(row=tmp_i + 12, column=2, value="+".join(i))
        tmp_j = 3
        for item in all_item:
            # 求平均值
            sum = 0
            for zh in i:
                if int(zh) <= 5:
                    if item[int(zh) - 1] is not None:
                        sum = sum + item[int(zh) - 1]
                    else:
                        sum = 0
                        break
                else:
                    if item[int(zh) - 4] is not None:
                        sum = sum + item[int(zh) - 4]
                    else:
                        sum = 0
                        break

            if sum != 0:
                ws.cell(row=tmp_i + 12, column=tmp_j, value=round(sum / no, 2))
            tmp_j = tmp_j + 1
        tmp_i = tmp_i + 1

    # 从第二行第一列开始写数据，第一行已经作为标题
    # i = 2
    # for item in all_item:
    #     ws.cell(row=i, column=1, value=str(item))
    #     i = i + 1
    wb.save(newfile)


# 程序入口
if __name__ == "__main__":
    # 3混1
    write_excel(3)
    write_excel(5)
    write_excel(7)
    write_excel(2)
