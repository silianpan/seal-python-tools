#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2020/10/19 9:48 上午
# @Author  : silianpan
# @Site    : 数据库数据补全
# @File    : db_fill.py
# @Software: PyCharm

# 1. 查询mysql数据库数据
# 2. 查询oracle数据库数据
# 3. 去重
# 4. 导出到excel

from openpyxl import load_workbook
from mysql_util import MysqlUtil
from oracle_util import OracleUtil


def bus_start(table_name):
    # 1. 查询oracle数据库数据
    ou = OracleUtil('192.168.1.200', 1521, 'xe', 'CDCZ_NPC2020MD', 'Asdf123')
    sql = "select t.chr_name from %s t order by t.chr_name" % ('ele_enterprise')
    rows = ou.select(sql)
    chr_names = []
    for row in rows:
        if row is not None:
            chr_name = str(row[0]).strip()
            if chr_name not in chr_names:
                chr_names.append(chr_name)

    # 2. 查询mysql数据库数据
    mu = MysqlUtil('192.168.1.200', 'bss_pro', 'root', 'Asdf@123')
    sql = "select t.gov_dept from %s t order by t.gov_dept" % (table_name)
    rows = mu.select(sql)
    ret_names = []
    for row in rows:
        gov_dept = str(row[0]).strip()
        if (gov_dept not in chr_names) and (gov_dept not in ret_names):
            ret_names.append(gov_dept)

    newfile = './ret_names.xlsx'
    wb = load_workbook(newfile)
    ws = wb['Sheet1']
    # 直接根据位置进行赋值
    ws['A1'] = '名称'

    i = 2
    for item in ret_names:
        ws.cell(row=i, column=1, value=str(item))
        i = i + 1
    wb.save(newfile)


def read_excel():
    excel_file = r'./ret_names_1.xlsx'
    inwb = load_workbook(excel_file)
    ws = inwb['Sheet1']

    # 获取sheet的最大行数和列数
    rows = ws.max_row
    cols = ws.max_column
    all_item = []
    for r in range(2, rows + 1):
        sname = ws.cell(r, 1).value
        tname = ws.cell(r, 2).value
        if sname and tname:
            item = {
                'sname': sname,
                'tname': tname
            }
            all_item.append(item)
    return all_item


def update_data(table_name):
    mu = MysqlUtil('192.168.1.200', 'bss_pro', 'root', 'Asdf@123')
    name_map = read_excel()
    for item in name_map:
        sql = "update %s t set t.gov_dept='%s' where t.gov_dept='%s'" % (table_name, item['tname'], item['sname'])
        print(sql)
        mu.update_sql(sql)


if __name__ == '__main__':
    bus_start('analysis_budget_dept_s2_out_general')
    # update_data('analysis_budget_dept_s8_summary_out')
