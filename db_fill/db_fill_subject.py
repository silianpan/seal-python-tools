#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2020/10/19 9:48 上午
# @Author  : silianpan
# @Site    : 数据库数据补全-预算科目
# @File    : db_fill.py
# @Software: PyCharm

# 1. 查询mysql数据库数据
# 2. 查询oracle数据库数据
# 3. 去重
# 4. 导出到excel

from openpyxl import load_workbook
from mysql_util import MysqlUtil
from oracle_util import OracleUtil


def bus_start(table_name):
    # 1. 查询oracle数据库数据
    ou = OracleUtil('192.168.1.200', 1521, 'xe', 'CDCZ_NPC2020MD', 'Asdf123')
    sql = "select distinct t.chr_name from %s t order by t.chr_name" % ('ele_budget_subject')
    rows = ou.select(sql)
    chr_names = []
    for row in rows:
        if row is not None:
            chr_name = str(row[0]).strip()
            if chr_name not in chr_names:
                chr_names.append(chr_name)

    # 2. 查询mysql数据库数据
    mu = MysqlUtil('192.168.1.200', 'bss_pro', 'root', 'Asdf@123')
    sql = "select distinct t.subject from %s t where t.year=2020 order by t.subject" % (table_name)
    rows = mu.select(sql)
    ret_names = []
    for row in rows:
        subject = str(row[0]).strip()
        if (subject not in chr_names) and (subject not in ret_names):
            ret_names.append(subject)

    newfile = './ret_names_subject.xlsx'
    wb = load_workbook(newfile)
    ws = wb['Sheet1']
    # 直接根据位置进行赋值
    ws['A1'] = '名称'

    i = 2
    for item in ret_names:
        ws.cell(row=i, column=1, value=str(item))
        i = i + 1
    wb.save(newfile)


def read_excel():
    excel_file = r'./ret_names_subject_1.xlsx'
    inwb = load_workbook(excel_file)
    ws = inwb['Sheet1']

    # 获取sheet的最大行数和列数
    rows = ws.max_row
    cols = ws.max_column
    all_item = []
    for r in range(2, rows + 1):
        sname = ws.cell(r, 1).value
        tname = ws.cell(r, 2).value
        if sname and tname:
            item = {
                'sname': str(sname).strip(),
                'tname': str(tname).strip()
            }
            all_item.append(item)
    return all_item


def update_data(table_name):
    mu = MysqlUtil('192.168.1.200', 'bss_pro', 'root', 'Asdf@123')
    name_map = read_excel()
    for item in name_map:
        sql = "update %s t set t.subject='%s' where t.year=2020 and t.subject='%s'" % (table_name, item['tname'], item['sname'])
        print(sql)
        mu.update_sql(sql)


def start(table_name):
    # update_data(table_name)
    bus_start(table_name)


if __name__ == '__main__':
    start('analysis_budget_dept_s1_out_fa')
